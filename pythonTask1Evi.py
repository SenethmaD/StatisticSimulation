import random
import math
import pandas as pd
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt

# Function to simulate dart throws and estimate pi
def throw_darts(num_darts):
    inside_circle = 0
    for _ in range(num_darts):
        x = random.random()
        y = random.random()
        distance = math.sqrt(x**2 + y**2)
        if distance <= 1:
            inside_circle += 1
    probability_of_hitting = inside_circle / num_darts
    estimated_pi = 4 * probability_of_hitting
    return estimated_pi

# Function to run multiple simulations with different parameters
def run_simulation():
    num_darts = int(input("\nEnter the number of darts to throw: "))
    num_experiments = 10
    num_runs = 10
    wb = Workbook()  # Create a new workbook
    for _ in range(num_experiments):
        pi_values = []
        for _ in range(num_runs):
            estimated_pi = throw_darts(num_darts)
            pi_values.append(estimated_pi)
        mean_pi = sum(pi_values) / num_runs
        mode_pi = max(set(pi_values), key=pi_values.count)
        experiment_results = {
            'Num Darts': num_darts,
            'Mean Pi': mean_pi,
            'Mode Pi': mode_pi,
        }
        # Create a new worksheet for each experiment
        ws = wb.create_sheet(title=f'Experiment {_ + 1}')
        ws.append(['Num Darts', 'Mean Pi', 'Mode Pi'])
        ws.append([experiment_results['Num Darts'], experiment_results['Mean Pi'], experiment_results['Mode Pi']])
    # Remove the default sheet created by openpyxl
    wb.remove(wb['Sheet'])
    # Save the workbook
    wb.save('dart_simulation_results_evidences.xlsx')

# Function to plot the results of the simulation and save the plots in a separate Excel sheet
def plot_results_and_save(dataframe, sheet_name, wb):
    plt.figure(figsize=(10, 6))
    plt.plot(dataframe['Num Darts'], dataframe['Mean Pi'], marker='o', label='Mean Pi')
    plt.plot(dataframe['Num Darts'], dataframe['Mode Pi'], marker='o', label='Mode Pi')
    plt.axhline(y=math.pi, color='r', linestyle='--', label='Exact Pi')
    plt.xscale('log')
    plt.xlabel('Number of Darts')
    plt.ylabel('Estimated Pi')
    plt.title('Estimation of Pi using Dart Simulation')
    plt.legend()
    plt.grid(True)
    
    # Save the plot as an image
    plot_image_path = f'plot_{sheet_name}.png'
    plt.savefig(plot_image_path)
    plt.close()  # Close the plot to free up memory

    # Add the plot image to the Excel file
    img = openpyxl.drawing.image.Image(plot_image_path)
    img.anchor = 'A1'  # Set the position of the image in the Excel sheet
    ws = wb.create_sheet(title=f'{sheet_name} Plot')
    ws.add_image(img)
    ws.column_dimensions['A'].width = 30  # Set the width of the first column to fit the plot title
    ws.append([f'Plot of {sheet_name}'])
    # Save the dataframe to the Excel file
    ws = wb.create_sheet(title=sheet_name)
    ws.append(['Num Darts', 'Mean Pi', 'Mode Pi'])
    for index, row in dataframe.iterrows():
        ws.append(row.tolist())

# Main block to define simulation parameters and execute the code
if __name__ == "__main__":
    # Run the simulation and store the results
    run_simulation()

    # Load the results from the Excel file and save the plots in a separate Excel sheet
    wb = openpyxl.load_workbook('dart_simulation_results_evidences.xlsx')
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('Experiment'):
            df = pd.read_excel('dart_simulation_results_evidences.xlsx', sheet_name=sheet_name)
            plot_results_and_save(df, sheet_name, wb)
    wb.save('dart_simulation_evidence.xlsx')

